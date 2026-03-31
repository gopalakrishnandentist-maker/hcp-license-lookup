"""
HCP License Lookup Tool — Search Preparation for Indian Medical Registries
===========================================================================
Parses Veeva OpenData HCP exports and generates structured search parameters
for license verification against NMC and State Medical Council registries.

Usage:
    python hcp_license_lookup.py --input hcp_export.xlsx --output lookup_prepared.xlsx

Reads Veeva HCP export, extracts relevant fields, maps each HCP to their
target State Medical Council, and generates ready-to-use search queries.
"""

import pandas as pd
import argparse
import os
import sys
import re
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# =============================================================================
# STATE MEDICAL COUNCIL REGISTRY
# =============================================================================

STATE_COUNCILS = {
    "Andhra Pradesh": ("AP Medical Council", "https://apmc.ap.gov.in/"),
    "Assam": ("Assam Medical Council", "https://assammedicalcouncil.org/"),
    "Bihar": ("Bihar Medical Council", "https://biharmedicalcouncil.org/"),
    "Chhattisgarh": ("CG Medical Council", "https://cghealth.nic.in/"),
    "Delhi": ("Delhi Medical Council", "https://delhimedicalcouncil.nic.in/"),
    "Goa": ("Goa Medical Council", "https://goamedicalcouncil.nic.in/"),
    "Gujarat": ("Gujarat Medical Council", "https://gmcgujarat.org/"),
    "Haryana": ("Haryana Medical Council", "https://haryanamedicalcouncil.nic.in/"),
    "Himachal Pradesh": ("HP Medical Council", "https://hpmedicalcouncil.nic.in/"),
    "Jharkhand": ("Jharkhand Medical Council", "https://jharkhandmedicalcouncil.nic.in/"),
    "Karnataka": ("Karnataka Medical Council", "https://kmc.karnataka.gov.in/"),
    "Kerala": ("Travancore Cochin Medical Council", "https://tcmc.kerala.gov.in/"),
    "Madhya Pradesh": ("MP Medical Council", "https://mpmedicalcouncil.org/"),
    "Maharashtra": ("Maharashtra Medical Council", "https://www.maharashtramedicalcouncil.in/"),
    "Odisha": ("Odisha Medical Council", "https://odishamedicalcouncil.org/"),
    "Punjab": ("Punjab Medical Council", "https://punjabmedicalcouncil.in/"),
    "Rajasthan": ("Rajasthan Medical Council", "https://rajmedicalcouncil.com/"),
    "Tamil Nadu": ("Tamil Nadu Medical Council", "https://tnmc.org/"),
    "Telangana": ("Telangana State Medical Council", "https://tsmc.telangana.gov.in/"),
    "Uttar Pradesh": ("UP State Medical Council", "https://upmedicalcouncil.org/"),
    "Uttarakhand": ("Uttarakhand Medical Council", "https://ukmc.uk.gov.in/"),
    "West Bengal": ("WB Medical Council", "https://wbmc.co.in/"),
}

CITY_TO_STATE = {
    "mumbai": "Maharashtra", "pune": "Maharashtra", "nagpur": "Maharashtra",
    "nashik": "Maharashtra", "thane": "Maharashtra", "aurangabad": "Maharashtra",
    "delhi": "Delhi", "new delhi": "Delhi",
    "gurgaon": "Haryana", "gurugram": "Haryana", "faridabad": "Haryana",
    "noida": "Uttar Pradesh", "greater noida": "Uttar Pradesh",
    "ghaziabad": "Uttar Pradesh", "lucknow": "Uttar Pradesh",
    "kanpur": "Uttar Pradesh", "varanasi": "Uttar Pradesh", "agra": "Uttar Pradesh",
    "chennai": "Tamil Nadu", "coimbatore": "Tamil Nadu", "madurai": "Tamil Nadu",
    "salem": "Tamil Nadu", "tiruchirappalli": "Tamil Nadu",
    "bangalore": "Karnataka", "bengaluru": "Karnataka", "mysore": "Karnataka",
    "mysuru": "Karnataka", "mangalore": "Karnataka", "mangaluru": "Karnataka",
    "hyderabad": "Telangana", "secunderabad": "Telangana",
    "kolkata": "West Bengal", "howrah": "West Bengal",
    "ahmedabad": "Gujarat", "surat": "Gujarat", "vadodara": "Gujarat",
    "rajkot": "Gujarat",
    "jaipur": "Rajasthan", "jodhpur": "Rajasthan", "udaipur": "Rajasthan",
    "kochi": "Kerala", "ernakulam": "Kerala", "trivandrum": "Kerala",
    "thiruvananthapuram": "Kerala", "calicut": "Kerala", "kozhikode": "Kerala",
    "chandigarh": "Punjab",
    "bhopal": "Madhya Pradesh", "indore": "Madhya Pradesh",
    "patna": "Bihar", "gaya": "Bihar",
    "bhubaneswar": "Odisha", "cuttack": "Odisha",
    "guwahati": "Assam",
    "ranchi": "Jharkhand", "jamshedpur": "Jharkhand",
    "dehradun": "Uttarakhand",
    "shimla": "Himachal Pradesh",
    "raipur": "Chhattisgarh",
    "panaji": "Goa", "margao": "Goa",
    "visakhapatnam": "Andhra Pradesh", "vijayawada": "Andhra Pradesh",
    "tirupati": "Andhra Pradesh",
}

# Veeva column name patterns
VEEVA_COLS = {
    "vid": ["hcp.vid__v", "hcp.vid__v (network id)", "hcp.vid__v (vid)", "vid"],
    "first_name": ["hcp.first_name__v", "hcp.first_name__v (first name)", "first_name"],
    "middle_name": ["hcp.middle_name__v", "hcp.middle_name__v (middle name)", "middle_name"],
    "last_name": ["hcp.last_name__v", "hcp.last_name__v (last name)", "last_name"],
    "intl_name": ["hcp.international_name__v", "hcp.international_name__v (international name)"],
    "specialty": ["hcp.specialty_1__v", "hcp.specialty_1__v (specialty 1)", "specialty_1"],
    "hco_name": ["hco.corporate_name__v", "hco.corporate_name__v (corporate name)"],
    "city": ["address.locality__v", "address.locality__v (city)", "city"],
    "state": ["address.administrative_area__v", "address.administrative_area__v (state/province)", "state"],
    "degree_cols": [f"hcp.custom_degree_{i}__c" for i in range(1, 6)],
}


def find_column(df, candidates):
    """Find the first matching column name (case-insensitive)."""
    df_cols_lower = {c.lower().strip(): c for c in df.columns}
    for candidate in candidates:
        if candidate.lower().strip() in df_cols_lower:
            return df_cols_lower[candidate.lower().strip()]
    return None


def get_value(row, col_name):
    """Safely get a string value from a row."""
    if col_name is None:
        return ""
    val = row.get(col_name, "")
    if pd.isna(val):
        return ""
    return str(val).strip()


def construct_name(row, col_intl, col_first, col_middle, col_last):
    """Build full HCP name from Veeva fields."""
    intl = get_value(row, col_intl)
    if intl:
        return intl

    parts = []
    for col in [col_first, col_middle, col_last]:
        v = get_value(row, col)
        if v:
            parts.append(v)
    return " ".join(parts)


def extract_affiliation(hco_raw):
    """Parse HCO corporate name into affiliation and department."""
    if not hco_raw:
        return "", ""

    dept = ""
    affiliation = hco_raw

    # Split department: "Hospital Name / Dept of X"
    if "/" in hco_raw:
        parts = hco_raw.split("/", 1)
        affiliation = parts[0].strip()
        dept = parts[1].strip()

    # Strip locality: "Hospital Name - City"
    if " - " in affiliation:
        affiliation = affiliation.split(" - ")[0].strip()

    return affiliation, dept


def resolve_state(city_raw, state_raw):
    """Determine state from explicit field or city mapping."""
    if state_raw:
        # Normalize common variations
        state_clean = state_raw.strip()
        for known_state in STATE_COUNCILS:
            if known_state.lower() == state_clean.lower():
                return known_state
        return state_clean

    if city_raw:
        city_lower = city_raw.lower().strip()
        # Try direct match
        if city_lower in CITY_TO_STATE:
            return CITY_TO_STATE[city_lower]
        # Try partial match
        for city_key, state_val in CITY_TO_STATE.items():
            if city_key in city_lower or city_lower in city_key:
                return state_val

    return ""


def get_council_info(state):
    """Look up State Medical Council name and URL."""
    if state in STATE_COUNCILS:
        return STATE_COUNCILS[state]
    # Case-insensitive fallback
    for key, val in STATE_COUNCILS.items():
        if key.lower() == state.lower():
            return val
    return ("Unknown — verify manually", "")


def build_search_queries(name, specialty, affiliation, city, state, council_name):
    """Generate structured search queries for registry lookup."""
    dr_name = f"Dr. {name}" if name else ""

    nmc_query = f'"{name}" "{council_name}" registration'
    if not name:
        nmc_query = ""

    web_query_parts = [f'"{dr_name}"']
    if specialty:
        web_query_parts.append(f'"{specialty}"')
    if affiliation:
        web_query_parts.append(f'"{affiliation}"')
    if city:
        web_query_parts.append(city)
    web_query = " ".join(web_query_parts)

    return nmc_query, web_query


def collect_degrees(row, df_columns):
    """Gather degree values from custom_degree columns."""
    degrees = []
    df_cols_lower = {c.lower().strip(): c for c in df_columns}

    for pattern in VEEVA_COLS["degree_cols"]:
        actual = df_cols_lower.get(pattern.lower().strip())
        if actual:
            val = get_value(row, actual)
            if val and val.lower() not in ("nan", "none", "0", ""):
                degrees.append(val)

    # Also check columns with "degree" in name
    for col in df_columns:
        if "degree" in col.lower() and col not in [df_cols_lower.get(p.lower().strip(), "") for p in VEEVA_COLS["degree_cols"]]:
            val = get_value(row, col)
            if val and val.lower() not in ("nan", "none", "0", ""):
                degrees.append(val)

    return ", ".join(degrees)


def style_output(path, df):
    """Apply professional formatting to the output Excel."""
    wb = load_workbook(path)
    ws = wb.active

    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(bold=True, color="FFFFFF", size=10)
    thin_border = Border(
        left=Side(style="thin", color="D9D9D9"),
        right=Side(style="thin", color="D9D9D9"),
        top=Side(style="thin", color="D9D9D9"),
        bottom=Side(style="thin", color="D9D9D9"),
    )

    # Style headers
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = thin_border

    # VID columns — force text format
    for col_idx in range(1, ws.max_column + 1):
        header_val = str(ws.cell(row=1, column=col_idx).value or "").lower()
        if "vid" in header_val:
            for row_idx in range(2, ws.max_row + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.number_format = "@"
                if cell.value is not None:
                    cell.value = str(cell.value)

    # Auto-fit columns
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            val = str(cell.value) if cell.value else ""
            max_len = max(max_len, len(val))
        ws.column_dimensions[col_letter].width = min(max_len + 3, 50)

    # Freeze header
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"

    wb.save(path)


def process(input_path, output_path):
    """Main processing pipeline."""
    print(f"Reading: {input_path}")

    # Read with all columns as string
    if input_path.lower().endswith(".csv"):
        df = pd.read_csv(input_path, dtype=str, encoding="utf-8")
    else:
        df = pd.read_excel(input_path, dtype=str)

    print(f"  Loaded {len(df):,} rows x {len(df.columns)} columns")

    # Resolve column names
    col_vid = find_column(df, VEEVA_COLS["vid"])
    col_first = find_column(df, VEEVA_COLS["first_name"])
    col_middle = find_column(df, VEEVA_COLS["middle_name"])
    col_last = find_column(df, VEEVA_COLS["last_name"])
    col_intl = find_column(df, VEEVA_COLS["intl_name"])
    col_spec = find_column(df, VEEVA_COLS["specialty"])
    col_hco = find_column(df, VEEVA_COLS["hco_name"])
    col_city = find_column(df, VEEVA_COLS["city"])
    col_state = find_column(df, VEEVA_COLS["state"])

    print(f"  VID column: {col_vid or 'NOT FOUND'}")
    print(f"  Name columns: intl={col_intl}, first={col_first}, last={col_last}")
    print(f"  Specialty: {col_spec or 'NOT FOUND'}")
    print(f"  HCO: {col_hco or 'NOT FOUND'}")
    print(f"  Location: city={col_city}, state={col_state}")

    # Process each row
    output_rows = []
    for idx, row in df.iterrows():
        row_num = idx + 1

        vid = get_value(row, col_vid)
        name = construct_name(row, col_intl, col_first, col_middle, col_last)
        specialty = get_value(row, col_spec)
        hco_raw = get_value(row, col_hco)
        city_raw = get_value(row, col_city)
        state_raw = get_value(row, col_state)

        affiliation, department = extract_affiliation(hco_raw)
        degrees = collect_degrees(row, df.columns)
        state = resolve_state(city_raw, state_raw)
        council_name, council_url = get_council_info(state)
        nmc_query, web_query = build_search_queries(
            name, specialty, affiliation, city_raw, state, council_name
        )

        output_rows.append({
            "Row_Num": row_num,
            "HCP_VID": vid,
            "HCP_Name": name,
            "Specialty": specialty,
            "Degrees": degrees,
            "Affiliation": affiliation,
            "Department": department,
            "City": city_raw,
            "State": state,
            "Target_Council": council_name,
            "Council_URL": council_url,
            "Search_Query_NMC": nmc_query,
            "Search_Query_Web": web_query,
            "License_Number": "",
            "Licensing_Council": "",
            "Verification_URL": "",
            "Confidence_Score": "",
            "Status": "Pending",
            "Notes": "",
        })

    out_df = pd.DataFrame(output_rows)

    # Write Excel
    out_df.to_excel(output_path, index=False, engine="openpyxl")
    style_output(output_path, out_df)

    # Summary
    states_covered = out_df["State"].replace("", pd.NA).dropna().nunique()
    councils_mapped = out_df[out_df["Target_Council"] != "Unknown — verify manually"]["Target_Council"].nunique()
    unmapped = (out_df["Target_Council"] == "Unknown — verify manually").sum()

    print(f"\n  Output: {output_path}")
    print(f"  Total HCPs: {len(out_df):,}")
    print(f"  States covered: {states_covered}")
    print(f"  Councils mapped: {councils_mapped}")
    if unmapped:
        print(f"  Unmapped (need manual council assignment): {unmapped}")
    print(f"\n  Done. Open the output file and fill in License_Number, Licensing_Council, and Status columns.")


def main():
    parser = argparse.ArgumentParser(
        description="HCP License Lookup — Search Preparation Tool",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
    python hcp_license_lookup.py --input hcp_export.xlsx --output lookup_prepared.xlsx
    python hcp_license_lookup.py --input hcp_data.csv --output lookup_prepared.xlsx
        """,
    )
    parser.add_argument("--input", required=True, help="Path to Veeva HCP export (CSV or Excel)")
    parser.add_argument("--output", required=True, help="Path for prepared lookup Excel output")

    args = parser.parse_args()

    if not os.path.isfile(args.input):
        print(f"ERROR: Input file not found: {args.input}")
        sys.exit(1)

    process(args.input, args.output)


if __name__ == "__main__":
    main()
